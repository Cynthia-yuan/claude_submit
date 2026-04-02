#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
命令对比HTML解析器
解析HTML格式的命令对比文档，提取命令信息并导出到Excel
"""

import os
import sys
import re
import argparse
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 复用 diff_parser.py 中的编码检测函数
def detect_file_encoding(file_path):
    """
    检测文件编码

    Args:
        file_path: 文件路径

    Returns:
        检测到的编码名称
    """
    encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'big5', 'utf-16', 'utf-16-le', 'utf-16-be', 'latin1']

    detected_encoding = None
    try:
        import chardet
        with open(file_path, 'rb') as f:
            raw_data = f.read(50000)
        result = chardet.detect(raw_data)
        if result and result['encoding']:
            detected_encoding = result['encoding'].lower()
            confidence = result.get('confidence', 0)
            print(f"检测到文件编码: {detected_encoding} (置信度: {confidence:.2f})")

            encoding_map = {
                'gb2312': 'gb18030',
                'gbk': 'gb18030',
                'big5': 'big5',
                'utf-8': 'utf-8',
                'utf-16': 'utf-16',
                'utf-16-le': 'utf-16-le',
                'utf-16-be': 'utf-16-be'
            }
            mapped_encoding = encoding_map.get(detected_encoding, detected_encoding)
            if mapped_encoding != detected_encoding:
                print(f"映射编码: {detected_encoding} -> {mapped_encoding}")

            if confidence > 0.5:
                encodings.insert(0, mapped_encoding)
    except ImportError:
        print("提示: 安装chardet可获得更准确的编码检测 (pip install chardet)")
    except Exception as e:
        print(f"编码检测失败: {e}")

    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                f.read(1000)
            print(f"使用编码: {encoding} (容错模式)")
            return encoding
        except Exception:
            continue

    print("警告: 无法确定编码，使用 latin1（可能导致乱码）")
    return 'latin1'


class CommandParser:
    """命令对比HTML解析器"""

    def __init__(self, html_file, output_file=None):
        """
        初始化解析器

        Args:
            html_file: HTML文件路径
            output_file: 输出Excel文件路径
        """
        self.html_file = html_file
        self.output_file = output_file or os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'output',
            'command_report.xlsx'
        )
        self.commands = []
        self.current_chapter = "未知章节"

    def find_chapter_for_element(self, element, soup):
        """
        查找元素所在章节

        Args:
            element: BeautifulSoup元素
            soup: 完整的BeautifulSoup对象

        Returns:
            章节名称
        """
        current = element
        last_heading = None

        for _ in range(20):
            if current is None:
                break

            prev = current.previous_sibling
            while prev:
                if prev.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p']:
                    text = prev.get_text(strip=True)
                    if text and (re.match(r'^\d+[\.、第]', text) or '章' in text or '节' in text):
                        return text
                prev = prev.previous_sibling

            if current.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                text = current.get_text(strip=True)
                if text:
                    return text

            current = current.parent

        element_index = list(soup.descendants).index(element) if element in list(soup.descendants) else 0

        all_headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        for heading in reversed(all_headings):
            heading_index = list(soup.descendants).index(heading) if heading in list(soup.descendants) else 0
            if heading_index < element_index:
                text = heading.get_text(strip=True)
                if text:
                    return text

        return "未知章节"

    def _parse_column_indices(self, headers):
        """
        解析列索引

        Args:
            headers: 表头列表

        Returns:
            列索引字典
        """
        indices = {
            '功能描述': None,
            '命令': None,
            '对比': None,
            '影响说明': None
        }

        for i, h in enumerate(headers):
            if '功能描述' in h:
                indices['功能描述'] = i
            elif '命令' in h:
                indices['命令'] = i
            elif '相比于' in h or '对比' in h:
                indices['对比'] = i
            elif '影响说明' in h:
                indices['影响说明'] = i

        return indices

    def _extract_command_row(self, row, col_idx, table, soup):
        """
        提取单行数据

        Args:
            row: 表格行元素
            col_idx: 列索引字典
            table: 表格元素
            soup: BeautifulSoup对象
        """
        cells = row.find_all(['td', 'th'])
        if not cells:
            return

        values = [cell.get_text(strip=True) for cell in cells]

        # 获取对比信息
        comparison_col = col_idx.get('对比')
        if comparison_col is None or comparison_col >= len(values):
            return

        comparison = values[comparison_col]

        # 跳过"无变化"或空值的行
        if not comparison or comparison == '无变化':
            return

        # 提取其他列数据
        function_desc = values[col_idx['功能描述']] if col_idx.get('功能描述') < len(values) else ''
        command = values[col_idx['命令']] if col_idx.get('命令') < len(values) else ''
        impact = values[col_idx['影响说明']] if col_idx.get('影响说明') and col_idx['影响说明'] < len(values) else ''

        # 获取章节
        chapter = self.find_chapter_for_element(table, soup)

        self.commands.append({
            'chapter': chapter,
            'function_desc': function_desc,
            'command': command,
            'comparison': comparison,
            'impact': impact,
            'verified': '待验证',
            'remark': ''
        })

    def extract_command_tables(self, soup):
        """
        提取命令对比格式的表格

        Args:
            soup: BeautifulSoup对象
        """
        print("正在查找命令对比表格...")

        tables = soup.find_all('table')
        print(f"  共找到 {len(tables)} 个表格")

        processed_count = 0

        for table in tables:
            # 检查表头
            headers = [th.get_text(strip=True) for th in table.find_all(['th', 'td'])]

            # 检测是否是命令对比格式
            has_function_desc = any('功能描述' in h for h in headers)
            has_command = any('命令' in h for h in headers)

            if has_function_desc and has_command:
                print(f"  找到命令对比表格，表头: {headers}")

                # 解析列索引
                col_idx = self._parse_column_indices(headers)

                # 验证必要的列存在
                if col_idx['功能描述'] is None or col_idx['命令'] is None or col_idx['对比'] is None:
                    print(f"  警告: 表格缺少必要列，跳过")
                    continue

                # 获取表格所在章节
                table_chapter = self.find_chapter_for_element(table, soup)
                print(f"  处理章节: {table_chapter}")

                # 解析数据行
                rows = table.find_all('tr')
                if len(rows) < 2:
                    continue

                skip_count = 0
                for row in rows[1:]:
                    # 快速检查：获取对比列的值
                    comparison_col = col_idx['对比']
                    cells = row.find_all(['td', 'th'])
                    if len(cells) > comparison_col:
                        comparison_text = cells[comparison_col].get_text(strip=True)
                        # 快速跳过"无变化"行
                        if not comparison_text or comparison_text == '无变化':
                            skip_count += 1
                            continue

                    self._extract_command_row(row, col_idx, table, soup)

                print(f"  提取了 {len(rows) - 1 - skip_count} 条记录（跳过 {skip_count} 条\"无变化\"记录）")
                processed_count += 1

        if processed_count == 0:
            print("  未找到命令对比格式的表格")
            print("  期望的表头格式: 功能描述 | 命令 | *(相比于)* | 影响说明")

    def export_to_excel(self):
        """导出到Excel"""
        if not self.commands:
            print("警告: 没有命令对比记录需要导出")
            return

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '命令对比'

        # 定义样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='673AB7', end_color='673AB7', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 表头
        headers = ['章节', '功能描述', '命令', '对比说明', '影响说明', '验证状态', '备注']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = alignment

        # 数据
        for row_num, cmd in enumerate(self.commands, 2):
            ws.cell(row=row_num, column=1, value=cmd['chapter'])
            ws.cell(row=row_num, column=2, value=cmd['function_desc'])
            ws.cell(row=row_num, column=3, value=cmd['command'])
            ws.cell(row=row_num, column=4, value=cmd['comparison'])
            ws.cell(row=row_num, column=5, value=cmd['impact'])
            ws.cell(row=row_num, column=6, value=cmd['verified'])
            ws.cell(row=row_num, column=7, value=cmd['remark'])

            # 应用样式
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = alignment

        # 调整列宽
        ws.column_dimensions['A'].width = 35  # 章节
        ws.column_dimensions['B'].width = 30  # 功能描述
        ws.column_dimensions['C'].width = 40  # 命令
        ws.column_dimensions['D'].width = 30  # 对比说明
        ws.column_dimensions['E'].width = 20  # 影响说明
        ws.column_dimensions['F'].width = 12  # 验证状态
        ws.column_dimensions['G'].width = 30  # 备注

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存文件
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        wb.save(self.output_file)

        print(f"\nExcel报告已生成: {self.output_file}")
        print(f"共提取 {len(self.commands)} 条命令对比记录")

    def generate_summary(self):
        """生成摘要"""
        print("\n命令对比摘要:")
        print(f"  总计: {len(self.commands)} 条")
        print(f"  待验证: {sum(1 for c in self.commands if c['verified'] == '待验证')} 条")

    def parse(self):
        """执行解析"""
        print(f"正在解析: {self.html_file}")

        # 加载HTML
        encoding = detect_file_encoding(self.html_file)
        try:
            with open(self.html_file, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read()
            soup = BeautifulSoup(content, 'html.parser')
            print(f"成功读取文件 ({len(content)} 字符)")
        except FileNotFoundError:
            print(f"错误: 文件不存在 - {self.html_file}")
            sys.exit(1)
        except Exception as e:
            print(f"错误: 加载HTML文件失败 - {e}")
            sys.exit(1)

        # 提取命令对比表格
        self.extract_command_tables(soup)

        if not self.commands:
            print("警告: 未找到任何命令对比记录")
            return

        # 生成摘要
        self.generate_summary()

        # 导出到Excel
        self.export_to_excel()


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='命令对比HTML解析器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 解析HTML文件
  python command_parser.py commands.html

  # 指定输出文件
  python command_parser.py commands.html -o output/report.xlsx
        """
    )
    parser.add_argument('html_file', help='HTML文件路径')
    parser.add_argument('-o', '--output', help='输出Excel文件路径')

    args = parser.parse_args()

    # 执行解析
    command_parser = CommandParser(args.html_file, args.output)
    command_parser.parse()


if __name__ == '__main__':
    main()
