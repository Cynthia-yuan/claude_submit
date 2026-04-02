#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
命令对比验证器
读取Excel中的命令对比记录，通过SSH连接到两个环境执行命令并验证
"""

import os
import sys
import json
import re
import argparse
import logging
from datetime import datetime
import paramiko
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class CommandValidator:
    """命令对比验证器"""

    def __init__(self, excel_file, env_old=None, env_new=None, config_file=None):
        """
        初始化验证器

        Args:
            excel_file: Excel文件路径
            env_old: 旧环境配置字典
            env_new: 新环境配置字典
            config_file: 配置文件路径
        """
        self.excel_file = excel_file
        self.config_file = config_file or os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'config.json'
        )
        self.log_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'output',
            'command_validation.log'
        )

        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)

        self.config = None
        self.env_old = env_old
        self.env_new = env_new
        self.commands = []
        self.validation_results = []

    def load_config(self):
        """加载配置文件"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            self.logger.info(f"配置文件加载成功: {self.config_file}")
        except FileNotFoundError:
            self.logger.warning(f"配置文件不存在: {self.config_file}")
            self.config = self.get_default_config()
        except json.JSONDecodeError as e:
            self.logger.warning(f"配置文件格式错误: {e}")
            self.config = self.get_default_config()

    def get_default_config(self):
        """获取默认配置"""
        return {
            "env_old": {
                "host": "",
                "port": 22,
                "username": "",
                "password": ""
            },
            "env_new": {
                "host": "",
                "port": 22,
                "username": "",
                "password": ""
            }
        }

    def get_env_config(self, env_name):
        """获取环境配置"""
        if env_name == 'env_old' and self.env_old:
            return self.env_old
        if env_name == 'env_new' and self.env_new:
            return self.env_new

        if self.config:
            return self.config.get(env_name, {})

        return {}

    def load_excel(self):
        """加载Excel文件"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb['命令对比']

            self.logger.info(f"正在读取Sheet页: 命令对比")

            # 跳过表头
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                self.commands.append({
                    'chapter': str(row[0]) if row[0] is not None else '',
                    'function_desc': str(row[1]) if len(row) > 1 and row[1] is not None else '',
                    'command': str(row[2]) if len(row) > 2 and row[2] is not None else '',
                    'comparison': str(row[3]) if len(row) > 3 and row[3] is not None else '',
                    'impact': str(row[4]) if len(row) > 4 and row[4] is not None else '',
                    'verified': str(row[5]) if len(row) > 5 and row[5] is not None else '待验证',
                    'remark': str(row[6]) if len(row) > 6 and row[6] is not None else ''
                })

            self.logger.info(f"从Excel加载了 {len(self.commands)} 条命令对比记录")
            return wb
        except FileNotFoundError:
            self.logger.error(f"Excel文件不存在: {self.excel_file}")
            sys.exit(1)
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {e}")
            sys.exit(1)

    def create_ssh_client(self, env_config):
        """创建SSH客户端连接"""
        if not env_config or not env_config.get('host'):
            self.logger.warning("环境配置不完整，跳过连接")
            return None

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try:
            password = env_config.get('password', '')

            if password:
                client.connect(
                    hostname=env_config['host'],
                    port=env_config.get('port', 22),
                    username=env_config['username'],
                    password=password,
                    timeout=10
                )
            else:
                self.logger.warning(f"未配置密码: {env_config['host']}")
                return None

            self.logger.info(f"SSH连接成功: {env_config['username']}@{env_config['host']}")
            return client
        except paramiko.AuthenticationException:
            self.logger.error(f"SSH认证失败: {env_config['host']}")
            return None
        except paramiko.SSHException as e:
            self.logger.error(f"SSH连接错误: {e}")
            return None
        except Exception as e:
            self.logger.error(f"连接失败: {e}")
            return None

    def execute_command(self, client, command):
        """执行SSH命令"""
        if not client:
            return -1, '', '未连接'

        try:
            stdin, stdout, stderr = client.exec_command(command)
            exit_code = stdout.channel.recv_exit_status()
            output = stdout.read().decode('utf-8', errors='ignore').strip()
            error = stderr.read().decode('utf-8', errors='ignore').strip()
            return exit_code, output, error
        except Exception as e:
            self.logger.error(f"执行命令失败: {command}, 错误: {e}")
            return -1, '', str(e)

    def _is_help_verification_needed(self, comparison):
        """
        判断是否需要使用 --help 验证

        Args:
            comparison: 对比描述

        Returns:
            (是否需要help验证, 提取的选项名)
        """
        if not comparison:
            return False, None

        # 匹配模式：新增选项、增加参数、新增参数等
        patterns = [
            r'新增\s*([\w-]+)\s*选项',
            r'增加\s*([\w-]+)\s*选项',
            r'新增\s*([\w-]+)\s*参数',
            r'增加\s*([\w-]+)\s*参数',
            r'添加\s*([\w-]+)\s*选项',
            r'revert\s*选项',
        ]

        for pattern in patterns:
            match = re.search(pattern, comparison)
            if match:
                option_name = match.group(1) if len(match.groups()) > 0 else 'revert'
                return True, option_name

        # 特殊处理：直接提到"revert"的情况
        if 'revert' in comparison.lower() and '选项' in comparison:
            return True, 'revert'

        return False, None

    def _extract_base_command(self, command):
        """
        从命令中提取基础命令（用于添加 --help）

        Args:
            command: 原始命令

        Returns:
            基础命令
        """
        # 简单提取：取第一个空格前的部分
        parts = command.split()
        if parts:
            base_cmd = parts[0]
            # 处理带路径的命令，如 /usr/bin/command
            if '/' in base_cmd:
                base_cmd = base_cmd.split('/')[-1]
            return base_cmd
        return command

    def validate_command_with_help(self, command, option_name, client_new):
        """
        使用 --help 验证选项是否存在

        Args:
            command: 原始命令
            option_name: 要验证的选项名
            client_new: 新环境SSH客户端

        Returns:
            (验证状态, 备注)
        """
        # 提取基础命令
        base_cmd = self._extract_base_command(command)

        # 构造 --help 命令
        help_command = f"{base_cmd} --help"

        self.logger.info(f"  使用 --help 验证选项: {option_name}")
        self.logger.info(f"  执行命令: {help_command}")

        exit_code, output, error = self.execute_command(client_new, help_command)

        # 组合输出和错误信息（有些命令的help在stderr中）
        full_output = (output + '\n' + error).lower()

        if exit_code != 0 and not output:
            return '失败', f'执行 {help_command} 失败: {error}'

        # 检查输出中是否包含选项名
        if option_name.lower() in full_output:
            return '通过', f'选项 {option_name} 存在于 --help 输出中 ✓'
        else:
            return '失败', f'选项 {option_name} 不存在于 --help 输出中\n输出: {output[:200]}...'

    def validate_command(self, cmd, client_old, client_new):
        """验证单个命令"""
        command = cmd['command']
        comparison = cmd['comparison']

        result = {
            'chapter': cmd['chapter'],
            'function_desc': cmd['function_desc'],
            'command': command,
            'comparison': comparison,
            'impact': cmd.get('impact', ''),
            'verified': '通过',
            'remark': ''
        }

        if not command:
            result['verified'] = '跳过'
            result['remark'] = '未提供命令'
            return result

        # 检查是否需要使用 --help 验证
        needs_help, option_name = self._is_help_verification_needed(comparison)

        if needs_help and option_name:
            # 使用 --help 验证
            verified, remark = self.validate_command_with_help(command, option_name, client_new)
            result['verified'] = verified
            result['remark'] = remark
            return result

        # 原有的验证逻辑：执行原命令
        self.logger.info(f"  执行命令: {command}")
        new_exit, new_output, new_err = self.execute_command(client_new, command)

        # 在旧环境执行命令
        old_exit, old_output, old_err = self.execute_command(client_old, command)

        # 验证逻辑
        if new_exit != 0:
            result['verified'] = '失败'
            result['remark'] = f'命令执行失败: {new_err or new_output}'
        elif '无变化' in comparison:
            # 期望无变化
            if old_exit != 0:
                result['verified'] = '警告'
                result['remark'] = f'旧环境命令执行失败，无法验证是否一致'
            elif old_output == new_output:
                result['verified'] = '通过'
                result['remark'] = f'新旧环境输出一致 ✓'
            else:
                result['verified'] = '失败'
                result['remark'] = f'期望无变化，但新旧环境输出不一致\n旧环境: {old_output[:100]}...\n新环境: {new_output[:100]}...'
        else:
            # 有变化 - 验证描述是否合理
            # 这里可以添加更复杂的验证逻辑
            # 当前简化版本：检查命令是否成功执行
            if old_exit != 0:
                # 旧环境失败，新环境成功
                result['verified'] = '通过'
                result['remark'] = f'新环境执行成功（旧环境失败）✓\n输出: {new_output[:100]}...'
            else:
                # 两个环境都成功
                result['verified'] = '部分验证'
                result['remark'] = f'新旧环境输出有变化（需人工确认是否符合描述）\n旧环境: {old_output[:100]}...\n新环境: {new_output[:100]}...'

        self.logger.info(f"  验证结果: {result['verified']}")
        return result

    def validate(self):
        """执行验证"""
        self.logger.info("=" * 60)
        self.logger.info("开始验证命令对比")
        self.logger.info(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info("=" * 60)

        # 加载配置
        self.load_config()

        # 获取环境配置
        env_old_config = self.get_env_config('env_old')
        env_new_config = self.get_env_config('env_new')

        # 打印环境信息
        self.logger.info(f"\n旧环境: {env_old_config.get('host', '未配置')} ({env_old_config.get('username', 'N/A')})")
        self.logger.info(f"新环境: {env_new_config.get('host', '未配置')} ({env_new_config.get('username', 'N/A')})")

        # 加载Excel
        wb = self.load_excel()

        # 如果没有有效的环境配置，跳过验证
        if not env_old_config.get('host') and not env_new_config.get('host'):
            self.logger.warning("\n环境配置不完整，跳过SSH验证")
            self.logger.warning("请使用以下方式配置环境:")
            self.logger.warning("  1. 命令行参数: --old-host / --new-host")
            self.logger.warning("  2. 配置文件: config.json")
            return

        # 创建SSH连接
        self.logger.info("\n正在连接SSH环境...")
        client_old = self.create_ssh_client(env_old_config)
        client_new = self.create_ssh_client(env_new_config)

        if not client_old:
            self.logger.warning("无法连接到旧环境，部分验证将无法执行")
        if not client_new:
            self.logger.warning("无法连接到新环境，部分验证将无法执行")

        # 验证每个命令
        self.logger.info("\n开始验证命令...")
        self.validation_results = []

        for i, cmd in enumerate(self.commands, 1):
            self.logger.info(f"\n[{i}/{len(self.commands)}] {cmd['function_desc']}")
            result = self.validate_command(cmd, client_old, client_new)
            self.validation_results.append(result)

        # 关闭SSH连接
        if client_old:
            client_old.close()
        if client_new:
            client_new.close()

        # 更新Excel文件
        self.update_excel(wb)

        # 生成HTML报告
        self.generate_html()

        # 生成验证摘要
        self.generate_summary()

        self.logger.info("\n" + "=" * 60)
        self.logger.info("验证完成")
        self.logger.info(f"日志文件: {self.log_file}")
        self.logger.info("=" * 60)

    def update_excel(self, wb):
        """更新Excel文件的验证结果"""
        fill_pass = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        fill_fail = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        fill_warn = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

        ws = wb['命令对比']
        row_num = 2

        for result in self.validation_results:
            for row in ws.iter_rows(min_row=row_num, values_only=False):
                if row[0].value == result['chapter'] and row[2].value == result['command']:
                    row[5].value = result['verified']  # 验证状态列
                    row[6].value = result['remark']  # 备注列

                    if result['verified'] == '通过':
                        row[5].fill = fill_pass
                    elif result['verified'] == '失败':
                        row[5].fill = fill_fail
                    elif result['verified'] == '警告':
                        row[5].fill = fill_warn
                    break
            row_num += 1

        wb.save(self.excel_file)
        self.logger.info(f"\nExcel文件已更新: {self.excel_file}")

    def generate_html(self):
        """生成HTML报告"""
        base_dir = os.path.dirname(os.path.abspath(self.excel_file))
        base_name = os.path.splitext(os.path.basename(self.excel_file))[0]
        self.html_file = os.path.join(base_dir, f"{base_name}.html")

        html_parts = ['<!DOCTYPE html>']
        html_parts.append('<html lang="zh-CN">')
        html_parts.append('<head>')
        html_parts.append('<meta charset="UTF-8">')
        html_parts.append('<meta name="viewport" content="width=device-width, initial-scale=1.0">')
        html_parts.append(f'<title>{base_name} - 验证报告</title>')
        html_parts.append('<style>')
        html_parts.append(self.get_html_css())
        html_parts.append('</style>')
        html_parts.append('</head>')
        html_parts.append('<body>')

        # 头部
        html_parts.append('<div class="container">')
        html_parts.append('<header>')
        html_parts.append(f'<h1>{base_name}</h1>')
        html_parts.append(f'<p class="timestamp">生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>')
        html_parts.append('</header>')

        # 总计汇总
        total_summary = self.calculate_summary()
        html_parts.append('<div class="total-summary">')
        html_parts.append(f'<span class="stat-item">总计: <strong>{total_summary["total"]}</strong></span>')
        if total_summary['passed'] > 0:
            html_parts.append(f'<span class="stat-item stat-pass">通过: {total_summary["passed"]}</span>')
        if total_summary['failed'] > 0:
            html_parts.append(f'<span class="stat-item stat-fail">失败: {total_summary["failed"]}</span>')
        if total_summary['warned'] > 0:
            html_parts.append(f'<span class="stat-item stat-warn">警告: {total_summary["warned"]}</span>')
        if total_summary['skipped'] > 0:
            html_parts.append(f'<span class="stat-item stat-skip">跳过: {total_summary["skipped"]}</span>')
        html_parts.append('</div>')

        # 筛选按钮
        html_parts.append('<div class="filter-bar">')
        html_parts.append('<button class="filter-btn active" data-filter="全部" onclick="filterTable(\'全部\')">全部</button>')
        html_parts.append(f'<button class="filter-btn" data-filter="通过" onclick="filterTable(\'通过\')">通过({total_summary["passed"]})</button>')
        html_parts.append(f'<button class="filter-btn" data-filter="失败" onclick="filterTable(\'失败\')">失败({total_summary["failed"]})</button>')
        html_parts.append(f'<button class="filter-btn" data-filter="警告" onclick="filterTable(\'警告\')">警告({total_summary["warned"]})</button>')
        html_parts.append(f'<button class="filter-btn" data-filter="跳过" onclick="filterTable(\'跳过\')">跳过({total_summary["skipped"]})</button>')
        html_parts.append('</div>')

        # 表格
        html_parts.append('<div class="content">')
        html_parts.append('<div class="table-wrapper">')
        html_parts.append('<table class="data-table">')
        html_parts.append('<colgroup><col style="width:15%"><col style="width:15%"><col style="width:25%"><col style="width:15%"><col style="width:10%"><col style="width:10%"><col style="width:10%"></colgroup>')
        html_parts.append('<thead><tr>')
        html_parts.append('<th>章节</th><th>功能描述</th><th>命令</th><th>对比说明</th><th>影响说明</th><th>验证状态</th><th>备注</th>')
        html_parts.append('</tr></thead>')
        html_parts.append('<tbody>')

        for result in self.validation_results:
            status_class = f' status-{result["verified"].lower()}' if result['verified'] in ['通过', '失败', '警告', '跳过'] else ''
            verified_status = result['verified']

            html_parts.append(f'<tr data-status="{verified_status}">')
            html_parts.append(f'<td>{result["chapter"]}</td>')
            html_parts.append(f'<td>{result["function_desc"]}</td>')
            html_parts.append(f'<td style="font-family:monospace;font-size:13px;">{result["command"]}</td>')
            html_parts.append(f'<td>{result["comparison"]}</td>')
            html_parts.append(f'<td>{result.get("impact", "")}</td>')
            html_parts.append(f'<td class="status{status_class}">{result["verified"]}</td>')
            html_parts.append(f'<td style="white-space:pre-wrap;">{result["remark"]}</td>')
            html_parts.append(f'</tr>')

        html_parts.append('</tbody>')
        html_parts.append('</table>')
        html_parts.append('</div>')
        html_parts.append('</div>')
        html_parts.append('</div>')

        # JavaScript
        html_parts.append('<script>')
        html_parts.append(self.get_html_js())
        html_parts.append('</script>')

        html_parts.append('</body>')
        html_parts.append('</html>')

        # 保存HTML文件
        html_content = '\n'.join(html_parts)
        try:
            with open(self.html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            self.logger.info(f"HTML报告已生成: {self.html_file}")
        except Exception as e:
            self.logger.error(f"生成HTML文件失败: {e}")

    def calculate_summary(self):
        """计算摘要"""
        summary = {'total': len(self.validation_results), 'passed': 0, 'failed': 0, 'warned': 0, 'skipped': 0}
        for result in self.validation_results:
            if result['verified'] == '通过':
                summary['passed'] += 1
            elif result['verified'] == '失败':
                summary['failed'] += 1
            elif result['verified'] == '警告':
                summary['warned'] += 1
            elif result['verified'] == '跳过':
                summary['skipped'] += 1
        return summary

    def generate_summary(self):
        """生成验证摘要"""
        summary = self.calculate_summary()

        self.logger.info("\n验证摘要:")
        self.logger.info(f"  总计: {summary['total']}")
        self.logger.info(f"  通过: {summary['passed']}")
        self.logger.info(f"  失败: {summary['failed']}")
        self.logger.info(f"  警告: {summary['warned']}")
        self.logger.info(f"  跳过: {summary['skipped']}")

    def get_html_css(self):
        """获取HTML CSS样式"""
        return '''
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    min-height: 100vh;
    padding: 20px;
}
.container {
    max-width: 1400px;
    margin: 0 auto;
    background: white;
    border-radius: 12px;
    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
    overflow: hidden;
    width: 95%;
}
header {
    background: linear-gradient(135deg, #673AB7 0%, #512DA8 100%);
    color: white;
    padding: 30px;
    text-align: center;
}
header h1 { font-size: 28px; margin-bottom: 10px; }
.timestamp { opacity: 0.9; font-size: 14px; }
.filter-bar {
    display: flex;
    justify-content: center;
    gap: 10px;
    padding: 20px;
    background: #f8f9fa;
    flex-wrap: wrap;
}
.filter-btn {
    padding: 8px 16px;
    border: 2px solid transparent;
    background: white;
    border-radius: 20px;
    font-size: 13px;
    cursor: pointer;
    transition: all 0.3s;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    font-weight: 500;
}
.filter-btn:hover {
    transform: translateY(-1px);
    box-shadow: 0 2px 5px rgba(0,0,0,0.15);
    border-color: #673AB7;
}
.filter-btn.active {
    background: linear-gradient(135deg, #673AB7 0%, #512DA8 100%);
    color: white;
    border-color: #673AB7;
}
.total-summary {
    display: flex;
    justify-content: center;
    gap: 20px;
    padding: 20px;
    background: #f8f9fa;
    flex-wrap: wrap;
}
.stat-item {
    padding: 10px 20px;
    background: white;
    border-radius: 25px;
    font-size: 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}
.stat-item strong { color: #673AB7; font-size: 18px; }
.stat-pass { color: #28a745; }
.stat-fail { color: #dc3545; }
.stat-warn { color: #ffc107; }
.stat-skip { color: #6c757d; }
.content { padding: 30px; overflow-x: auto; }
.table-wrapper { overflow-x: auto; -webkit-overflow-scrolling: touch; }
.data-table { width: 100%; border-collapse: collapse; box-shadow: 0 2px 8px rgba(0,0,0,0.1); table-layout: fixed; }
.data-table thead { background: linear-gradient(135deg, #673AB7 0%, #512DA8 100%); color: white; }
.data-table th { padding: 15px; text-align: left; font-weight: 600; white-space: nowrap; }
.data-table td { padding: 12px 15px; border-bottom: 1px solid #e9ecef; word-wrap: break-word; }
.data-table tbody tr:hover { background: #f8f9fa; }
.status-pass { color: #28a745; font-weight: 600; }
.status-fail { color: #dc3545; font-weight: 600; }
.status-warn { color: #ffc107; font-weight: 600; }
.status-skip { color: #6c757d; font-weight: 600; }
@media (max-width: 768px) {
    .data-table { font-size: 12px; }
    .data-table th, .data-table td { padding: 8px; }
}
'''

    def get_html_js(self):
        """获取HTML JavaScript代码"""
        return '''
function filterTable(status) {
    document.querySelectorAll('.filter-btn').forEach(btn => {
        btn.classList.remove('active');
        if (btn.dataset.filter === status) {
            btn.classList.add('active');
        }
    });

    document.querySelectorAll('.data-table tbody tr').forEach(row => {
        if (status === '全部') {
            row.style.display = '';
        } else {
            const rowStatus = row.dataset.status || '待验证';
            if (rowStatus === status) {
                row.style.display = '';
            } else {
                row.style.display = 'none';
            }
        }
    });
}
document.addEventListener('DOMContentLoaded', function() {
    document.querySelector('.filter-btn')?.click();
});
'''


def parse_env_config(args, prefix):
    """解析环境配置"""
    DEFAULT_USER = 'root'
    DEFAULT_PASS = 'Huawei12#$'

    config = {
        'host': getattr(args, f'{prefix}_host', None) or '',
        'port': getattr(args, f'{prefix}_port', 22),
        'username': getattr(args, f'{prefix}_user', None) or DEFAULT_USER,
        'password': getattr(args, f'{prefix}_pass', None) or DEFAULT_PASS
    }

    if not config['host']:
        return None

    return config


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='命令对比验证器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python command_validator.py command_report.xlsx \\
    --old-host 192.168.1.10 --new-host 192.168.1.20
        """
    )

    parser.add_argument('excel_file', help='Excel命令对比报告文件路径')
    parser.add_argument('-c', '--config', help='配置文件路径')

    # 旧环境参数
    parser.add_argument('--old-host', help='旧环境主机地址')
    parser.add_argument('--old-port', type=int, default=22, help='旧环境SSH端口')
    parser.add_argument('--old-user', default='root', help='旧环境用户名')
    parser.add_argument('--old-pass', default='Huawei12#$', help='旧环境密码')

    # 新环境参数
    parser.add_argument('--new-host', help='新环境主机地址')
    parser.add_argument('--new-port', type=int, default=22, help='新环境SSH端口')
    parser.add_argument('--new-user', default='root', help='新环境用户名')
    parser.add_argument('--new-pass', default='Huawei12#$', help='新环境密码')

    args = parser.parse_args()

    # 解析环境配置
    env_old = parse_env_config(args, 'old')
    env_new = parse_env_config(args, 'new')

    # 执行验证
    validator = CommandValidator(args.excel_file, env_old, env_new, args.config)
    validator.validate()


if __name__ == '__main__':
    main()
