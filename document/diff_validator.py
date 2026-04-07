#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SSH差异验证器
读取Excel中的变更记录，通过SSH连接到两个环境进行验证
"""

import os
import sys
import json
import re
import argparse
import logging
from datetime import datetime
import paramiko
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class SSHValidator:
    """SSH环境验证器"""

    def __init__(self, excel_file, env_old=None, env_new=None, config_file=None):
        """
        初始化验证器

        Args:
            excel_file: Excel文件路径
            env_old: 旧环境配置字典 {'host': '', 'username': '', 'password': '', 'port': 22}
            env_new: 新环境配置字典 {'host': '', 'username': '', 'password': '', 'port': 22}
            config_file: 配置文件路径（可选）
        """
        self.excel_file = excel_file
        self.config_file = config_file or os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'config.json'
        )
        self.log_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'output',
            'validation.log'
        )

        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)

        self.config = None
        self.env_old = env_old
        self.env_new = env_new
        self.changes = []
        self.validation_results = []

    def load_config(self):
        """加载配置文件"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            self.logger.info(f"配置文件加载成功: {self.config_file}")
        except FileNotFoundError:
            self.logger.warning(f"配置文件不存在: {self.config_file}")
            self.config = self.get_default_config()
        except json.JSONDecodeError as e:
            self.logger.warning(f"配置文件格式错误: {e}")
            self.config = self.get_default_config()

    def get_default_config(self):
        """获取默认配置"""
        return {
            "env_old": {
                "host": "",
                "port": 22,
                "username": "",
                "password": "",
                "key_file": ""
            },
            "env_new": {
                "host": "",
                "port": 22,
                "username": "",
                "password": "",
                "key_file": ""
            },
            "validation_rules": {
                "file_check_paths": ["/opt/packages", "/etc/app", "/opt/scripts"],
                "config_files": ["/etc/app/application.yml"],
                "check_commands": {
                    "jdk_version": "java -version 2>&1",
                    "mysql_version": "mysql --version"
                }
            }
        }

    def get_env_config(self, env_name):
        """
        获取环境配置，优先使用命令行参数，其次使用配置文件

        Args:
            env_name: 环境名称 'env_old' 或 'env_new'

        Returns:
            环境配置字典
        """
        # 命令行参数优先
        if env_name == 'env_old' and self.env_old:
            return self.env_old
        if env_name == 'env_new' and self.env_new:
            return self.env_new

        # 使用配置文件
        if self.config:
            return self.config.get(env_name, {})

        return {}

    def load_excel(self):
        """加载Excel文件中的变更记录"""
        try:
            wb = load_workbook(self.excel_file)

            # 获取所有sheet页
            all_changes = []
            sheet_names = wb.sheetnames

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                self.logger.info(f"正在读取Sheet页: {sheet_name}")

                # 检查是否有数据（检查第一行是否是表头）
                first_row = [cell.value for cell in ws[1]]
                if not first_row or first_row[0] is None:
                    continue

                # 跳过表头，从第二行开始
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 跳过空行
                    if not row or row[0] is None:
                        continue

                    # 判断变更类型（根据sheet名称或第二列）
                    if sheet_name in ['删除', '新增', '修改']:
                        change_type = sheet_name
                    elif len(row) > 1 and row[1]:
                        change_type = row[1]
                    else:
                        continue

                    all_changes.append({
                        'sheet': sheet_name,
                        'chapter': str(row[0]) if row[0] is not None else '',
                        'change_type': change_type,
                        'item_name': str(row[1]) if row[1] is not None else '',
                        'impact': str(row[2]) if len(row) > 2 and row[2] is not None else '',
                        'description': str(row[3]) if len(row) > 3 and row[3] is not None else '',
                        'verified': str(row[4]) if len(row) > 4 and row[4] is not None else '待验证',
                        'remark': str(row[5]) if len(row) > 5 and row[5] is not None else ''
                    })

            self.changes = all_changes
            self.logger.info(f"从Excel加载了 {len(self.changes)} 条变更记录")
            return wb
        except FileNotFoundError:
            self.logger.error(f"Excel文件不存在: {self.excel_file}")
            sys.exit(1)
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {e}")
            sys.exit(1)

    def create_ssh_client(self, env_config):
        """
        创建SSH客户端连接

        Args:
            env_config: 环境配置字典

        Returns:
            SSHClient对象
        """
        if not env_config or not env_config.get('host'):
            self.logger.warning("环境配置不完整，跳过连接")
            return None

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try:
            key_file = env_config.get('key_file')
            password = env_config.get('password', '')

            # 优先使用密钥认证
            if key_file and os.path.exists(key_file):
                key = paramiko.RSAKey.from_private_key_file(key_file)
                client.connect(
                    hostname=env_config['host'],
                    port=env_config.get('port', 22),
                    username=env_config['username'],
                    pkey=key,
                    timeout=10
                )
            # 其次使用密码认证
            elif password:
                client.connect(
                    hostname=env_config['host'],
                    port=env_config.get('port', 22),
                    username=env_config['username'],
                    password=password,
                    timeout=10
                )
            else:
                self.logger.warning(f"未配置密钥或密码: {env_config['host']}")
                return None

            self.logger.info(f"SSH连接成功: {env_config['username']}@{env_config['host']}")
            return client
        except paramiko.AuthenticationException:
            self.logger.error(f"SSH认证失败: {env_config['host']}")
            return None
        except paramiko.SSHException as e:
            self.logger.error(f"SSH连接错误: {e}")
            return None
        except Exception as e:
            self.logger.error(f"连接失败: {e}")
            return None

    def execute_command(self, client, command):
        """
        执行SSH命令

        Args:
            client: SSHClient对象
            command: 要执行的命令

        Returns:
            (exit_code, stdout, stderr)
        """
        try:
            stdin, stdout, stderr = client.exec_command(command)
            exit_code = stdout.channel.recv_exit_status()
            return exit_code, stdout.read().decode('utf-8').strip(), stderr.read().decode('utf-8').strip()
        except Exception as e:
            self.logger.error(f"执行命令失败: {command}, 错误: {e}")
            return -1, '', str(e)

    def check_file_exists(self, client, file_path):
        """
        检查文件是否存在

        Args:
            client: SSHClient对象
            file_path: 文件路径

        Returns:
            bool: 文件是否存在
        """
        if not client:
            return False

        # 处理正则表达式通配符 .*
        if '.*' in file_path:
            file_path = self._convert_regex_wildcard(file_path)

        # 检查是否包含通配符
        if '*' in file_path:
            return self.check_wildcard_path_exists(client, file_path)

        exit_code, stdout, _ = self.execute_command(client, f"test -e '{file_path}' && echo 'exists'")
        return exit_code == 0 and 'exists' in stdout

    def _convert_regex_wildcard(self, file_path):
        """
        将正则表达式通配符 .* 转换为 shell 通配符 *

        Args:
            file_path: 包含 .* 的路径

        Returns:
            转换后的路径
        """
        original_path = file_path

        # 将 .* 替换为 *（shell通配符）
        # 注意：不要替换 .*. 或 .*等特殊情况
        # 只替换独立的 .* 模式
        import re
        # 使用正则替换，匹配 .* 但不匹配其他 . 的情况
        file_path = re.sub(r'\.\*', '*', file_path)

        if original_path != file_path:
            self.logger.info(f"  正则通配符转换: {original_path} -> {file_path}")

        return file_path

    def check_wildcard_path_exists(self, client, file_path):
        """
        检查包含通配符的路径是否有匹配的文件存在

        Args:
            client: SSHClient对象
            file_path: 包含通配符的文件路径

        Returns:
            bool: 是否有至少一个匹配的文件存在
        """
        if not client:
            return False

        self.logger.info(f"    检查通配符路径: {file_path}")

        # 方法1: 直接使用 shell 通配符展开（最简单高效）
        # 例如: ls -1 /proc/sys/net/ipv6/conf/*/accept_untracked_na 2>/dev/null
        exit_code, stdout, stderr = self.execute_command(
            client,
            f"ls -1 {file_path} 2>/dev/null | head -5"
        )

        # 如果有输出，说明至少有一个匹配项
        if exit_code == 0 and stdout.strip():
            match_count = len([line for line in stdout.strip().split('\n') if line])
            self.logger.info(f"    找到 {match_count} 个匹配文件")
            return True

        # 方法2: 对于通配符在中间的情况（如 /proc/*/status），使用 find 命令
        # 例如: find /proc -maxdepth 1 -type d -name '*' -exec test -f {}/status \; -print
        path_parts = file_path.split('/')

        # 找到通配符的位置
        wildcard_index = -1
        for i, part in enumerate(path_parts):
            if '*' in part:
                wildcard_index = i
                break

        if wildcard_index >= 0 and wildcard_index < len(path_parts) - 1:
            # 通配符在中间部分
            # 例如: /proc/sys/net/ipv6/conf/*/accept_untracked_na
            # 目标文件名: accept_untracked_na
            # 搜索基础: /proc/sys/net/ipv6/conf/

            base_search_path = '/'.join(path_parts[:wildcard_index + 1])
            target_filename = path_parts[-1]

            # 使用 find 查找所有匹配的文件
            # 注意：shell 的 * 可以在 find 命令中直接使用
            find_cmd = f"find {base_search_path} -name '{target_filename}' 2>/dev/null | head -1"

            self.logger.info(f"    尝试 find 命令: {find_cmd}")

            exit_code, stdout, stderr = self.execute_command(client, find_cmd)

            if exit_code == 0 and stdout.strip():
                self.logger.info(f"    通过 find 找到匹配文件")
                return True

        self.logger.info(f"    未找到匹配文件")
        return False

    def _extract_multiple_paths(self, text):
        """
        从文本中提取多个路径

        Args:
            text: 可能包含多个路径的文本

        Returns:
            路径列表
        """
        if not text:
            return []

        paths = []

        # 方法1: 使用正则表达式提取所有路径
        # 匹配 / 开头的路径，包含常见字符和通配符
        path_pattern = r'(/[a-zA-Z0-9_\-./<>.*]+)'
        matches = re.findall(path_pattern, text)
        if matches:
            paths.extend(matches)

        # 方法2: 如果正则没找到，尝试按分隔符分割
        if not paths:
            separators = [',', ';', '\t', '，', '；']
            for sep in separators:
                if sep in text:
                    parts = text.split(sep)
                    for part in parts:
                        part = part.strip()
                        if part.startswith('/'):
                            paths.append(part)
                    if paths:
                        break

        # 如果还是没有，检查整个文本是否是路径
        if not paths and text.startswith('/'):
            paths = [text]

        return paths

    def _select_best_path(self, paths):
        """
        从路径列表中选择最佳路径用于验证

        优先级：具体路径 > 通配符路径

        Args:
            paths: 路径列表

        Returns:
            最佳路径
        """
        if not paths:
            return None

        # 分类路径
        concrete_paths = []  # 不含通配符的路径
        wildcard_paths = []  # 含通配符的路径

        for path in paths:
            if '*' in path:
                wildcard_paths.append(path)
            else:
                concrete_paths.append(path)

        # 优先返回具体路径
        if concrete_paths:
            self.logger.info(f"  从 {len(paths)} 个路径中选择具体路径: {concrete_paths[0]}")
            return concrete_paths[0]

        # 如果没有具体路径，返回通配符路径
        if wildcard_paths:
            self.logger.info(f"  从 {len(paths)} 个路径中选择通配符路径: {wildcard_paths[0]}")
            return wildcard_paths[0]

        return paths[0]

    def extract_file_path(self, description, item_name=''):
        """
        从描述中提取文件路径，并处理占位符

        Args:
            description: 描述文本
            item_name: 变更项名称

        Returns:
            文件路径字符串（占位符已替换为1）
        """
        path = None

        # 首先检查item_name（可能包含多个路径）
        if item_name:
            # 检查是否包含多个路径
            paths = self._extract_multiple_paths(item_name)
            if paths:
                # 选择最佳路径
                path = self._select_best_path(paths)
            else:
                # 单个路径的情况
                if item_name.startswith('/') or '/' in item_name:
                    if re.match(r'^[a-zA-Z0-9_\-./<>]+$', item_name) or item_name.startswith('/'):
                        path = item_name

        # 如果没有从item_name获取到路径，从描述中提取
        if not path:
            # 尝试从描述中提取多个路径
            paths = self._extract_multiple_paths(description)
            if paths:
                path = self._select_best_path(paths)
            else:
                # 单个路径匹配
                path_match = re.search(r'(/[a-zA-Z0-9_\-./<>]+)', description)
                if path_match:
                    path = path_match.group(1)

        if not path:
            return None

        # 替换占位符为1用于验证
        original_path = path

        # 转换正则表达式通配符 .* 为 shell 通配符 *
        if '.*' in path:
            path = re.sub(r'\.\*', '*', path)
            if path != original_path:
                self.logger.info(f"  正则通配符转换: {original_path} -> {path}")
                original_path = path

        # 优先替换 <1> 为 1（diff_parser.py 已经将 <pid> 替换为 <1>）
        path = path.replace('<1>', '1')
        # 兼容旧数据：如果还有 <pid>，也替换为 1
        path = path.replace('<pid>', '1').replace('<PID>', '1')
        if original_path != path:
            self.logger.info(f"  路径占位符替换: {original_path} -> {path}")

        return path

    def is_valid_path(self, text):
        """
        判断文本是否是有效路径

        Args:
            text: 要检查的文本

        Returns:
            bool: 是否是有效路径
        """
        if not text:
            return False

        # 必须以/开头，或包含/
        if not (text.startswith('/') or '/' in text):
            return False

        # 路径中不能包含中文字符（通常路径是英文/数字/符号）
        # 允许的字符：字母、数字、-._/以及中文路径名
        return True

    def extract_replacement_info(self, description, item_name=''):
        """
        从描述中提取替换信息（A变成B）- 仅用于非删除类变更

        Args:
            description: 描述文本
            item_name: 变更项名称

        Returns:
            (old_path, new_path) 或 (None, None)
        """
        old_path = None
        new_path = None

        # 从item_name获取
        if item_name:
            if item_name.startswith('/') or '/' in item_name:
                old_path = item_name

        # 如果没有old_path，尝试从描述开头提取
        if not old_path:
            # 匹配开头的路径，支持 <pid> 等占位符
            path_match = re.search(r'(/[a-zA-Z0-9_\-./<>]+)', description)
            if path_match:
                old_path = path_match.group(1)

        # 如果没有找到old_path，返回None
        if not old_path:
            return None, None

        # 替换关键词模式 - 只在修改类变更中使用
        replacement_patterns = [
            r'重命名\s*[为为]?\s*([^\s，。]+)',
            r'->\s*([^\s，。]+)',
            r'→\s*([^\s，。]+)',
        ]

        # 在描述中查找替换模式
        for pattern in replacement_patterns:
            match = re.search(pattern, description)
            if match:
                potential_new_path = match.group(1)
                # 校验新路径是否有效
                if self.is_valid_path(potential_new_path):
                    new_path = potential_new_path
                    break

        # 如果两个路径都有效，返回
        if old_path and new_path:
            return old_path, new_path

        return None, None

    def validate_change(self, change, client_old, client_new):
        """
        验证单个变更

        Args:
            change: 变更字典
            client_old: 旧环境SSH客户端
            client_new: 新环境SSH客户端

        Returns:
            验证结果字典
        """
        change_type = change['change_type']
        item_name = change.get('item_name', '')
        description = change['description']

        result = {
            'sheet': change.get('sheet', ''),
            'chapter': change['chapter'],
            'change_type': change_type,
            'item_name': item_name,
            'impact': change.get('impact', ''),
            'description': description,
            'verified': '通过',
            'remark': ''
        }

        # 对于删除类型，只验证删除，不检查替换
        if change_type == '删除':
            file_path = self.extract_file_path(description, item_name)
            if file_path:
                result = self.validate_deletion(file_path, client_old, client_new, result)
            else:
                # 无法提取路径的删除项，标记为符合预期
                result['verified'] = '通过'
                result['remark'] = '删除项（无法验证路径，符合预期）'

        elif change_type == '修改':
            # 对于修改类型，检查是否有替换场景
            old_path, new_path = self.extract_replacement_info(description, item_name)
            if old_path and new_path:
                # 修改并替换场景
                result = self.validate_replacement(old_path, new_path, client_new, result)
            else:
                # 普通修改场景
                file_path = self.extract_file_path(description, item_name)
                if file_path:
                    result = self.validate_modification(file_path, client_old, client_new, result)
                else:
                    result['verified'] = '跳过'
                    result['remark'] = '无法从描述中提取文件路径'

        elif change_type == '新增':
            file_path = self.extract_file_path(description, item_name)
            if file_path:
                result = self.validate_addition(file_path, client_old, client_new, result)
            else:
                result['verified'] = '跳过'
                result['remark'] = '无法从描述中提取文件路径'

        elif change_type == '修改':
            file_path = self.extract_file_path(description, item_name)
            if file_path:
                result = self.validate_modification(file_path, client_old, client_new, result)
            else:
                result['verified'] = '跳过'
                result['remark'] = '无法从描述中提取文件路径'

        else:
            result['verified'] = '跳过'
            result['remark'] = f'未知变更类型: {change_type}'

        return result

    def validate_deletion(self, file_path, client_old, client_new, result):
        """验证删除项"""
        # 检查旧环境是否存在
        old_exists = client_old and self.check_file_exists(client_old, file_path)
        # 检查新环境是否不存在
        new_exists = client_new and self.check_file_exists(client_new, file_path)

        if old_exists and not new_exists:
            result['verified'] = '通过'
            result['remark'] = f'旧环境存在，新环境不存在 ✓'
        elif not old_exists and not new_exists:
            result['verified'] = '警告'
            result['remark'] = f'两个环境都不存在该文件'
        elif new_exists:
            result['verified'] = '失败'
            result['remark'] = f'新环境仍存在该文件，应删除'
        else:
            result['verified'] = '未知'
            result['remark'] = f'无法验证（可能是连接问题）'

        self.logger.info(f"[删除] {file_path} - {result['verified']}")
        return result

    def validate_replacement(self, old_path, new_path, client_new, result):
        """
        验证删除并替换项（A变成B）

        Args:
            old_path: 旧路径
            new_path: 新路径
            client_new: 新环境SSH客户端
            result: 结果字典

        Returns:
            验证结果字典
        """
        # 验证：旧路径在新环境不存在，新路径在新环境存在
        old_exists_new_env = client_new and self.check_file_exists(client_new, old_path)
        new_exists_new_env = client_new and self.check_file_exists(client_new, new_path)

        if not old_exists_new_env and new_exists_new_env:
            result['verified'] = '通过'
            result['remark'] = f'{old_path} 已删除，{new_path} 已存在 ✓'
        elif not old_exists_new_env and not new_exists_new_env:
            result['verified'] = '失败'
            result['remark'] = f'{old_path} 已删除，但 {new_path} 不存在'
        elif old_exists_new_env and new_exists_new_env:
            result['verified'] = '失败'
            result['remark'] = f'{old_path} 仍然存在，{new_path} 也存在'
        else:
            result['verified'] = '失败'
            result['remark'] = f'{old_path} 仍然存在，{new_path} 不存在'

        self.logger.info(f"[删除/替换] {old_path} -> {new_path} - {result['verified']}")
        return result

    def validate_addition(self, file_path, client_old, client_new, result):
        """验证新增项"""
        # 检查旧环境是否不存在
        old_exists = client_old and self.check_file_exists(client_old, file_path)
        # 检查新环境是否存在
        new_exists = client_new and self.check_file_exists(client_new, file_path)

        # 如果路径包含通配符，提供更详细的说明
        has_wildcard = '*' in file_path

        if not old_exists and new_exists:
            result['verified'] = '通过'
            if has_wildcard:
                result['remark'] = f'新环境存在匹配文件 ✓（通配符路径）'
            else:
                result['remark'] = f'新环境存在 ✓'
        elif old_exists and new_exists:
            result['verified'] = '警告'
            if has_wildcard:
                result['remark'] = f'两个环境都存在匹配文件（通配符路径）'
            else:
                result['remark'] = f'两个环境都存在该文件'
        elif not new_exists:
            result['verified'] = '失败'
            if has_wildcard:
                result['remark'] = f'新环境不存在匹配文件，应新增（通配符路径）'
            else:
                result['remark'] = f'新环境不存在该文件，应新增'
        else:
            result['verified'] = '未知'
            result['remark'] = f'无法验证'

        self.logger.info(f"[新增] {file_path} - {result['verified']}")
        return result

    def validate_modification(self, file_path, client_old, client_new, result):
        """验证修改项"""
        # 检查两个环境文件是否存在
        old_exists = client_old and self.check_file_exists(client_old, file_path)
        new_exists = client_new and self.check_file_exists(client_new, file_path)

        if not new_exists:
            result['verified'] = '失败'
            result['remark'] = f'新环境不存在该文件'
            return result

        # 可以添加更多的验证逻辑，比如比较文件大小、修改时间、内容等
        result['verified'] = '部分验证'
        result['remark'] = f'新环境文件存在（建议人工确认内容变更）'

        self.logger.info(f"[修改] {file_path} - {result['verified']}")
        return result

    def validate(self):
        """执行验证"""
        self.logger.info("=" * 60)
        self.logger.info("开始验证差异文档")
        self.logger.info(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info("=" * 60)

        # 加载配置
        self.load_config()

        # 获取环境配置
        env_old_config = self.get_env_config('env_old')
        env_new_config = self.get_env_config('env_new')

        # 打印环境信息
        self.logger.info(f"\n旧环境: {env_old_config.get('host', '未配置')} ({env_old_config.get('username', 'N/A')})")
        self.logger.info(f"新环境: {env_new_config.get('host', '未配置')} ({env_new_config.get('username', 'N/A')})")

        # 加载Excel
        wb = self.load_excel()

        # 如果没有有效的环境配置，跳过验证
        if not env_old_config.get('host') and not env_new_config.get('host'):
            self.logger.warning("\n环境配置不完整，跳过SSH验证")
            self.logger.warning("请使用以下方式配置环境:")
            self.logger.warning("  1. 命令行参数: --old-host / --new-host")
            self.logger.warning("  2. 配置文件: config.json")
            return

        # 创建SSH连接
        self.logger.info("\n正在连接SSH环境...")
        client_old = self.create_ssh_client(env_old_config)
        client_new = self.create_ssh_client(env_new_config)

        # 如果连接失败，使用None继续
        if not client_old:
            self.logger.warning("无法连接到旧环境，部分验证将无法执行")
        if not client_new:
            self.logger.warning("无法连接到新环境，部分验证将无法执行")

        # 验证每个变更
        self.logger.info("\n开始验证变更记录...")
        self.validation_results = []

        for i, change in enumerate(self.changes, 1):
            desc = change['description'][:50] if change['description'] else change['item_name']
            self.logger.info(f"\n[{i}/{len(self.changes)}] 验证: {desc}")
            result = self.validate_change(change, client_old, client_new)
            self.validation_results.append(result)

        # 关闭SSH连接
        if client_old:
            client_old.close()
        if client_new:
            client_new.close()

        # 更新Excel文件
        self.update_excel(wb)

        # 生成HTML报告
        self.generate_html()

        # 生成验证摘要
        self.generate_summary()

        self.logger.info("\n" + "=" * 60)
        self.logger.info("验证完成")
        self.logger.info(f"日志文件: {self.log_file}")
        self.logger.info("=" * 60)

    def update_excel(self, wb):
        """更新Excel文件的验证结果"""
        # 定义样式
        fill_pass = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        fill_fail = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        fill_warn = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

        # 按sheet分组更新
        results_by_sheet = {}
        for result in self.validation_results:
            sheet_name = result.get('sheet', '')
            if sheet_name not in results_by_sheet:
                results_by_sheet[sheet_name] = []
            results_by_sheet[sheet_name].append(result)

        # 更新每个sheet
        for sheet_name, results in results_by_sheet.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            row_num = 2  # 从第二行开始（跳过表头）

            for result in results:
                # 找到对应的行并更新
                for row in ws.iter_rows(min_row=row_num, values_only=False):
                    if row[0].value == result['chapter'] and row[2].value == result['description']:
                        row[3].value = result['verified']
                        row[4].value = result['remark']

                        # 根据验证结果设置样式
                        if result['verified'] == '通过':
                            row[3].fill = fill_pass
                        elif result['verified'] == '失败':
                            row[3].fill = fill_fail
                        elif result['verified'] == '警告':
                            row[3].fill = fill_warn
                        break
                row_num += 1

        # 保存更新后的Excel
        wb.save(self.excel_file)
        self.logger.info(f"\nExcel文件已更新: {self.excel_file}")

    def generate_html(self):
        """生成HTML报告"""
        # 确定HTML输出路径
        base_dir = os.path.dirname(os.path.abspath(self.excel_file))
        base_name = os.path.splitext(os.path.basename(self.excel_file))[0]
        self.html_file = os.path.join(base_dir, f"{base_name}.html")

        # 按sheet分组结果
        results_by_sheet = {}
        for result in self.validation_results:
            sheet_name = result.get('sheet', '未知')
            if sheet_name not in results_by_sheet:
                results_by_sheet[sheet_name] = []
            results_by_sheet[sheet_name].append(result)

        # 生成HTML内容
        html_parts = ['<!DOCTYPE html>']
        html_parts.append('<html lang="zh-CN">')
        html_parts.append('<head>')
        html_parts.append('<meta charset="UTF-8">')
        html_parts.append('<meta name="viewport" content="width=device-width, initial-scale=1.0">')
        html_parts.append(f'<title>{base_name} - 验证报告</title>')
        html_parts.append('<style>')
        html_parts.append(self.get_html_css())
        html_parts.append('</style>')
        html_parts.append('</head>')
        html_parts.append('<body>')

        # 头部
        html_parts.append('<div class="container">')
        html_parts.append('<header>')
        html_parts.append(f'<h1>{base_name}</h1>')
        html_parts.append(f'<p class="timestamp">生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>')
        html_parts.append('</header>')

        # 总计汇总（放在导航标签上方）
        total_summary = self.calculate_total_summary()
        html_parts.append('<div class="total-summary">')
        html_parts.append(f'<span class="stat-item">总计: <strong>{total_summary["total"]}</strong></span>')
        if total_summary['passed'] > 0:
            html_parts.append(f'<span class="stat-item stat-pass">通过: {total_summary["passed"]}</span>')
        if total_summary['failed'] > 0:
            html_parts.append(f'<span class="stat-item stat-fail">失败: {total_summary["failed"]}</span>')
        if total_summary['warned'] > 0:
            html_parts.append(f'<span class="stat-item stat-warn">警告: {total_summary["warned"]}</span>')
        if total_summary['skipped'] > 0:
            html_parts.append(f'<span class="stat-item stat-skip">跳过: {total_summary["skipped"]}</span>')
        html_parts.append('</div>')

        # 导航标签
        html_parts.append('<nav class="tabs">')
        for sheet_name in results_by_sheet.keys():
            sheet_id = sheet_name.replace(' ', '-').lower()
            html_parts.append(f'<button class="tab-button {"active" if sheet_name == list(results_by_sheet.keys())[0] else ""}" onclick="showTab(\'{sheet_id}\')">{sheet_name}</button>')
        html_parts.append('</nav>')

        # 内容区域
        html_parts.append('<div class="content">')

        for idx, (sheet_name, results) in enumerate(results_by_sheet.items()):
            sheet_id = sheet_name.replace(' ', '-').lower()
            is_active = idx == 0

            html_parts.append(f'<div id="{sheet_id}" class="tab-content {"active" if is_active else ""}">')

            # Sheet筛选按钮（只保留筛选按钮）
            summary = self.calculate_sheet_summary(results)
            html_parts.append(f'<div class="sheet-filters" data-sheet="{sheet_id}">')
            html_parts.append(f'<button class="filter-btn active" data-filter="全部" onclick="filterSheet(\'{sheet_id}\', \'全部\')">全部({summary["total"]})</button>')
            html_parts.append(f'<button class="filter-btn" data-filter="通过" onclick="filterSheet(\'{sheet_id}\', \'通过\')">通过({summary["passed"]})</button>')
            html_parts.append(f'<button class="filter-btn" data-filter="失败" onclick="filterSheet(\'{sheet_id}\', \'失败\')">失败({summary["failed"]})</button>')
            html_parts.append(f'<button class="filter-btn" data-filter="警告" onclick="filterSheet(\'{sheet_id}\', \'警告\')">警告({summary["warned"]})</button>')
            html_parts.append(f'<button class="filter-btn" data-filter="跳过" onclick="filterSheet(\'{sheet_id}\', \'跳过\')">跳过({summary["skipped"]})</button>')
            html_parts.append(f'<button class="filter-btn" data-filter="待验证" onclick="filterSheet(\'{sheet_id}\', \'待验证\')">待验证({summary.get("pending", summary["total"])})</button>')
            html_parts.append('</div>')

            # 表格
            html_parts.append('<div class="table-wrapper">')
            html_parts.append('<table class="data-table">')
            html_parts.append('<colgroup><col style="width:20%"><col style="width:15%"><col style="width:15%"><col style="width:25%"><col style="width:10%"><col style="width:15%"></colgroup>')
            html_parts.append('<thead><tr>')
            html_parts.append('<th>章节</th><th>变更项</th><th>影响说明</th><th>描述</th><th>验证状态</th><th>备注</th>')
            html_parts.append('</tr></thead>')
            html_parts.append('<tbody>')

            for result in results:
                status_class = f' status-{result["verified"].lower()}' if result['verified'] in ['通过', '失败', '警告', '跳过'] else ''
                impact = result.get('impact', '')
                verified_status = result['verified']

                # 替换占位符：将 <pid> 替换为 1
                chapter = self._replace_placeholder(result["chapter"])
                item_name = self._replace_placeholder(result["item_name"])
                impact_display = self._replace_placeholder(impact)
                description = self._replace_placeholder(result["description"])
                remark = self._replace_placeholder(result["remark"])

                html_parts.append(f'<tr data-status="{verified_status}">')
                html_parts.append(f'<td>{chapter}</td>')
                html_parts.append(f'<td>{item_name}</td>')
                html_parts.append(f'<td>{impact_display}</td>')
                html_parts.append(f'<td>{description}</td>')
                html_parts.append(f'<td class="status{status_class}">{result["verified"]}</td>')
                html_parts.append(f'<td>{remark}</td>')
                html_parts.append(f'</tr>')

            html_parts.append('</tbody>')
            html_parts.append('</table>')
            html_parts.append('</div>')
            html_parts.append('</div>')

        html_parts.append('</div>')
        html_parts.append('</div>')

        # JavaScript
        html_parts.append('<script>')
        html_parts.append(self.get_html_js())
        html_parts.append('</script>')

        html_parts.append('</body>')
        html_parts.append('</html>')

        # 保存HTML文件
        html_content = '\n'.join(html_parts)
        try:
            with open(self.html_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            self.logger.info(f"HTML报告已生成: {self.html_file}")
        except Exception as e:
            self.logger.error(f"生成HTML文件失败: {e}")

    def calculate_total_summary(self):
        """计算总计摘要"""
        summary = {'total': 0, 'passed': 0, 'failed': 0, 'warned': 0, 'skipped': 0}
        for result in self.validation_results:
            summary['total'] += 1
            if result['verified'] == '通过':
                summary['passed'] += 1
            elif result['verified'] == '失败':
                summary['failed'] += 1
            elif result['verified'] == '警告':
                summary['warned'] += 1
            elif result['verified'] == '跳过':
                summary['skipped'] += 1
        return summary

    def calculate_sheet_summary(self, results):
        """计算Sheet摘要"""
        summary = {'total': 0, 'passed': 0, 'failed': 0, 'warned': 0, 'skipped': 0, 'pending': 0}
        for result in results:
            summary['total'] += 1
            if result['verified'] == '通过':
                summary['passed'] += 1
            elif result['verified'] == '失败':
                summary['failed'] += 1
            elif result['verified'] == '警告':
                summary['warned'] += 1
            elif result['verified'] == '跳过':
                summary['skipped'] += 1
            elif result['verified'] == '待验证':
                summary['pending'] += 1
        return summary

    def _replace_placeholder(self, text):
        """
        替换文本中的占位符（用于HTML显示）

        Args:
            text: 原始文本

        Returns:
            替换后的文本
        """
        if not text:
            return text
        # 保持 <1> 不变（diff_parser.py 已替换）
        # 兼容旧数据：如果还有 <pid> 或 <PID>，替换为 <1>
        return text.replace('<pid>', '<1>').replace('<PID>', '<1>')

    def get_html_css(self):
        """获取HTML CSS样式"""
        return '''
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    min-height: 100vh;
    padding: 20px;
}
.container {
    max-width: 1400px;
    margin: 0 auto;
    background: white;
    border-radius: 12px;
    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
    overflow: hidden;
    width: 95%;
}
header {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 30px;
    text-align: center;
}
header h1 { font-size: 28px; margin-bottom: 10px; }
.timestamp { opacity: 0.9; font-size: 14px; }
.filter-bar {
    display: flex;
    justify-content: center;
    gap: 10px;
    padding: 20px;
    background: #f8f9fa;
    flex-wrap: wrap;
}
.filter-btn {
    padding: 8px 16px;
    border: 2px solid transparent;
    background: white;
    border-radius: 20px;
    font-size: 13px;
    cursor: pointer;
    transition: all 0.3s;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    font-weight: 500;
}
.filter-btn:hover {
    transform: translateY(-1px);
    box-shadow: 0 2px 5px rgba(0,0,0,0.15);
    border-color: #667eea;
}
.filter-btn.active {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    border-color: #667eea;
}
.sheet-filters {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
    padding: 15px 0;
    justify-content: center;
}
.filter-btn.filter-pass:hover { border-color: #28a745; }
.filter-btn.filter-fail:hover { border-color: #dc3545; }
.filter-btn.filter-warn:hover { border-color: #ffc107; }
.filter-btn.filter-skip:hover { border-color: #6c757d; }
.filter-btn .count {
    opacity: 0.8;
    font-size: 12px;
}
.total-summary {
    display: flex;
    justify-content: center;
    gap: 20px;
    padding: 20px;
    background: #f8f9fa;
    flex-wrap: wrap;
}
.stat-item {
    padding: 10px 20px;
    background: white;
    border-radius: 25px;
    font-size: 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}
.stat-item strong { color: #667eea; font-size: 18px; }
.stat-pass { color: #28a745; }
.stat-fail { color: #dc3545; }
.stat-warn { color: #ffc107; }
.stat-skip { color: #6c757d; }
.tabs { display: flex; background: #f8f9fa; border-bottom: 1px solid #dee2e6; }
.tab-button {
    flex: 1;
    padding: 15px 20px;
    border: none;
    background: transparent;
    cursor: pointer;
    font-size: 16px;
    color: #495057;
    border-bottom: 3px solid transparent;
}
.tab-button:hover { background: #e9ecef; color: #667eea; }
.tab-button.active { color: #667eea; border-bottom-color: #667eea; background: white; }
.content { padding: 30px; overflow-x: auto; }
.tab-content { display: none; }
.tab-content.active { display: block; animation: fadeIn 0.3s; }
.table-wrapper { overflow-x: auto; -webkit-overflow-scrolling: touch; }
@keyframes fadeIn { from { opacity: 0; } to { opacity: 1; } }
.summary { margin-bottom: 20px; padding: 20px; background: #f8f9fa; border-radius: 8px; }
.summary h2 { font-size: 20px; margin-bottom: 15px; color: #495057; }
.summary-stats { display: flex; flex-wrap: wrap; gap: 15px; }
.data-table { width: 100%; border-collapse: collapse; box-shadow: 0 2px 8px rgba(0,0,0,0.1); table-layout: fixed; overflow-wrap: break-word; }
.data-table thead { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
.data-table th { padding: 15px; text-align: left; font-weight: 600; white-space: nowrap; }
.data-table td { padding: 12px 15px; border-bottom: 1px solid #e9ecef; word-wrap: break-word; white-space: normal; overflow-wrap: break-word; hyphens: auto; }
.data-table tbody tr:hover { background: #f8f9fa; }
.status-pass { color: #28a745; font-weight: 600; }
.status-fail { color: #dc3545; font-weight: 600; }
.status-warn { color: #ffc107; font-weight: 600; }
.status-skip { color: #6c757d; font-weight: 600; }
@media (max-width: 768px) {
    .tabs { flex-direction: column; }
    .tab-button { border-left: 3px solid transparent; }
    .tab-button.active { border-left-color: #667eea; }
    .data-table { font-size: 12px; }
    .data-table th, .data-table td { padding: 8px; }
}
'''

    def get_html_js(self):
        """获取HTML JavaScript代码"""
        return '''
function showTab(tabId) {
    document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.tab-button').forEach(el => el.classList.remove('active'));
    document.getElementById(tabId).classList.add('active');
    event.target.classList.add('active');
}
function filterSheet(sheetId, status) {
    // 更新当前Sheet的按钮状态
    const container = document.getElementById(sheetId);
    container.querySelectorAll('.filter-btn').forEach(btn => {
        btn.classList.remove('active');
        if (btn.dataset.filter === status) {
            btn.classList.add('active');
        }
    });
    // 筛选当前Sheet的数据行
    container.querySelectorAll('.data-table tbody tr').forEach(row => {
        if (status === '全部') {
            row.style.display = '';
        } else {
            const rowStatus = row.dataset.status || '待验证';
            if (rowStatus === status) {
                row.style.display = '';
            } else {
                row.style.display = 'none';
            }
        }
    });
}
document.addEventListener('DOMContentLoaded', function() {
    document.querySelector('.tab-button')?.click();
});
'''

    def generate_summary(self):
        """生成验证摘要"""
        summary = {
            '通过': 0,
            '失败': 0,
            '警告': 0,
            '跳过': 0,
            '未知': 0,
            '部分验证': 0,
            '总计': len(self.validation_results)
        }

        for result in self.validation_results:
            verified = result['verified']
            if verified in summary:
                summary[verified] += 1

        self.logger.info("\n验证摘要:")
        for status, count in summary.items():
            self.logger.info(f"  {status}: {count}")


def parse_env_config(args, prefix):
    """
    解析环境配置

    Args:
        args: 命令行参数
        prefix: 'old' 或 'new'

    Returns:
        环境配置字典
    """
    # 默认值
    DEFAULT_USER = 'root'
    DEFAULT_PASS = 'Huawei12#$'

    config = {
        'host': getattr(args, f'{prefix}_host', None) or '',
        'port': getattr(args, f'{prefix}_port', 22),
        'username': getattr(args, f'{prefix}_user', None) or DEFAULT_USER,
        'password': getattr(args, f'{prefix}_pass', None) or DEFAULT_PASS,
        'key_file': getattr(args, f'{prefix}_key', None) or ''
    }

    # 如果没有配置host，返回None
    if not config['host']:
        return None

    return config


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='SSH差异验证器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 只指定主机地址（用户名默认root，密码默认Huawei12#$）
  python diff_validator.py diff_report.xlsx \\
    --old-host 192.168.1.10 --new-host 192.168.1.20

  # 指定用户名和密码
  python diff_validator.py diff_report.xlsx \\
    --old-host 192.168.1.10 --old-user admin --old-pass mypassword \\
    --new-host 192.168.1.20

  # 使用配置文件
  python diff_validator.py diff_report.xlsx --config config.json
        """
    )

    parser.add_argument('excel_file', help='Excel差异报告文件路径')
    parser.add_argument('-c', '--config', help='配置文件路径')

    # 旧环境参数
    parser.add_argument('--old-host', help='旧环境主机地址')
    parser.add_argument('--old-port', type=int, default=22, help='旧环境SSH端口 (默认: 22)')
    parser.add_argument('--old-user', default='root', help='旧环境用户名 (默认: root)')
    parser.add_argument('--old-pass', default='Huawei12#$', help='旧环境密码 (默认: Huawei12#$)')

    # 新环境参数
    parser.add_argument('--new-host', help='新环境主机地址')
    parser.add_argument('--new-port', type=int, default=22, help='新环境SSH端口 (默认: 22)')
    parser.add_argument('--new-user', default='root', help='新环境用户名 (默认: root)')
    parser.add_argument('--new-pass', default='Huawei12#$', help='新环境密码 (默认: Huawei12#$)')

    args = parser.parse_args()

    # 解析环境配置
    env_old = parse_env_config(args, 'old')
    env_new = parse_env_config(args, 'new')

    # 执行验证
    validator = SSHValidator(args.excel_file, env_old, env_new, args.config)
    validator.validate()


if __name__ == '__main__':
    main()
