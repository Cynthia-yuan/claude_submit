#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HTML差异文档解析器
解析HTML格式的版本差异文档，提取变更信息并导出到Excel
"""

import os
import sys
import re
import argparse
from datetime import datetime
from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def detect_file_encoding(file_path):
    """
    检测文件编码

    Args:
        file_path: 文件路径

    Returns:
        检测到的编码名称
    """
    # 常见的中文编码（gb18030放在前面，因为它是gb2312/gbk的超集，兼容性更好）
    encodings = ['utf-8', 'gb18030', 'gbk', 'gb2312', 'big5', 'utf-16', 'utf-16-le', 'utf-16-be', 'latin1']

    # 先尝试用chardet检测
    detected_encoding = None
    try:
        import chardet
        with open(file_path, 'rb') as f:
            raw_data = f.read(50000)  # 读取更多字节提高检测准确性
        result = chardet.detect(raw_data)
        if result and result['encoding']:
            detected_encoding = result['encoding'].lower()
            confidence = result.get('confidence', 0)
            print(f"检测到文件编码: {detected_encoding} (置信度: {confidence:.2f})")

            # 将chardet检测到的编码映射到实际可用的编码
            encoding_map = {
                'gb2312': 'gb18030',  # gb18030兼容gb2312且更健壮
                'gbk': 'gb18030',     # gb18030兼容gbk且更健壮
                'big5': 'big5',
                'utf-8': 'utf-8',
                'utf-16': 'utf-16',
                'utf-16-le': 'utf-16-le',
                'utf-16-be': 'utf-16-be'
            }
            mapped_encoding = encoding_map.get(detected_encoding, detected_encoding)
            if mapped_encoding != detected_encoding:
                print(f"映射编码: {detected_encoding} -> {mapped_encoding}")

            # 优先使用检测到的编码
            if confidence > 0.5:
                encodings.insert(0, mapped_encoding)
    except ImportError:
        print("提示: 安装chardet可获得更准确的编码检测 (pip install chardet)")
    except Exception as e:
        print(f"编码检测失败: {e}")

    # 逐个尝试常见编码，使用容错模式
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                f.read(1000)  # 尝试读取一小部分
            print(f"使用编码: {encoding} (容错模式)")
            return encoding
        except Exception:
            continue

    # 如果都失败，使用latin1（它不会抛出解码错误）
    print("警告: 无法确定编码，使用 latin1（可能导致乱码）")
    return 'latin1'


class DiffParser:
    """HTML差异文档解析器"""

    def __init__(self, html_file, output_file=None, include_chapters=None, skip_chapters=None):
        """
        初始化解析器

        Args:
            html_file: HTML文件路径
            output_file: 输出Excel文件路径
            include_chapters: 包含的章节列表，如 ['5', '5.1', '5.2']，None表示全部
            skip_chapters: 跳过的章节列表，如 ['1', '2']，None表示不跳过
        """
        self.html_file = html_file
        self.output_file = output_file or os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            'output',
            'diff_report.xlsx'
        )
        self.changes = []
        self.current_chapter = "未知章节"
        self.include_chapters = self._parse_chapters(include_chapters)
        self.skip_chapters = self._parse_chapters(skip_chapters)

        # 打印配置信息
        if self.include_chapters:
            print(f"只检索章节: {self.include_chapters}")
        if self.skip_chapters:
            print(f"跳过章节: {self.skip_chapters}")

    def _parse_chapters(self, chapters_str):
        """
        解析章节字符串

        Args:
            chapters_str: 章节字符串，如 "5,5.1,5.2" 或 ['5', '5.1']

        Returns:
            章节集合
        """
        if not chapters_str:
            return None

        if isinstance(chapters_str, str):
            # 支持逗号、空格、分号分隔
            parts = re.split(r'[,;，\s]+', chapters_str.strip())
            return set(p for p in parts if p)
        elif isinstance(chapters_str, list):
            return set(chapters_str)
        return None

    def _should_include_chapter(self, chapter_name):
        """
        判断章节是否应该被包含

        Args:
            chapter_name: 章节名称

        Returns:
            True表示包含，False表示跳过
        """
        # 提取章节编号（如 "第5章 交付件差异" -> "5"）
        chapter_num = self._extract_chapter_number(chapter_name)

        # 如果指定了包含列表，只处理列表中的章节
        if self.include_chapters:
            # 检查章节编号是否在包含列表中
            for include in self.include_chapters:
                if chapter_num.startswith(include) or include in chapter_name:
                    # 再检查是否在跳过列表中
                    if self.skip_chapters:
                        for skip in self.skip_chapters:
                            if chapter_num.startswith(skip) or skip in chapter_name:
                                return False
                    return True
            return False

        # 如果没有指定包含列表，检查是否在跳过列表中
        if self.skip_chapters:
            for skip in self.skip_chapters:
                if chapter_num.startswith(skip) or skip in chapter_name:
                    return False

        return True

    def _extract_chapter_number(self, chapter_name):
        """
        从章节名称中提取章节编号

        Args:
            chapter_name: 章节名称，如 "第5章 交付件差异" 或 "5.1 安装包变更"

        Returns:
            章节编号，如 "5" 或 "5.1"
        """
        if not chapter_name:
            return ""

        # 匹配 "第X章" 格式
        match = re.search(r'第(\d+(\.\d+)*)章', chapter_name)
        if match:
            return match.group(1)

        # 匹配 "X.X" 或 "X" 开头的格式
        match = re.match(r'^(\d+(\.\d+)*)(\.|\s|、|$)', chapter_name)
        if match:
            return match.group(1)

        # 返回原始名称（用于模糊匹配）
        return chapter_name

    def load_html(self):
        """加载HTML文件"""
        try:
            # 自动检测文件编码
            encoding = detect_file_encoding(self.html_file)

            # 使用容错模式读取文件，忽略无法解码的字节
            with open(self.html_file, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read()

            print(f"成功读取文件 ({len(content)} 字符)")
            return BeautifulSoup(content, 'html.parser')
        except FileNotFoundError:
            print(f"错误: 文件不存在 - {self.html_file}")
            sys.exit(1)
        except Exception as e:
            print(f"错误: 加载HTML文件失败 - {e}")
            sys.exit(1)

    def find_chapter_for_element(self, element, soup):
        """
        查找元素所在章节

        Args:
            element: BeautifulSoup元素
            soup: 完整的BeautifulSoup对象

        Returns:
            章节名称
        """
        # 向上查找标题元素
        current = element
        last_heading = None

        # 向上遍历DOM树
        for _ in range(20):  # 限制遍历层级
            if current is None:
                break

            # 查找前面的兄弟元素中的标题
            prev = current.previous_sibling
            while prev:
                if prev.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p']:
                    text = prev.get_text(strip=True)
                    # 检查是否是章节标题（包含数字或特定格式）
                    if text and (re.match(r'^\d+[\.、第]', text) or '章' in text or '节' in text):
                        return text
                prev = prev.previous_sibling

            # 检查当前元素本身
            if current.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                text = current.get_text(strip=True)
                if text:
                    return text

            # 向父元素查找
            current = current.parent

        # 如果向上查找失败，尝试在整个文档中查找最近的标题
        # 获取元素在文档中的位置
        element_index = list(soup.descendants).index(element) if element in list(soup.descendants) else 0

        # 查找元素之前的所有标题
        all_headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        for heading in reversed(all_headings):
            heading_index = list(soup.descendants).index(heading) if heading in list(soup.descendants) else 0
            if heading_index < element_index:
                text = heading.get_text(strip=True)
                if text:
                    return text

        return "未知章节"

    def normalize_change_type(self, change_type_text):
        """
        标准化变更类型文本

        Args:
            change_type_text: 原始变更类型文本

        Returns:
            标准化的变更类型：删除/新增/修改
        """
        if not change_type_text:
            return '未知'

        text = change_type_text.strip().lower()

        # 删除类型关键词
        delete_keywords = ['删除', '减少', '移除', '废弃', '去掉', '删除掉', 'del', 'delete', 'remove']
        # 新增类型关键词
        add_keywords = ['新增', '增加', '添加', '新增了', 'add', 'new', 'create']
        # 修改类型关键词
        modify_keywords = ['修改', '变更', '更新', '改变', '调整', 'modify', 'change', 'update']

        for keyword in delete_keywords:
            if keyword in text:
                return '删除'

        for keyword in add_keywords:
            if keyword in text:
                return '新增'

        for keyword in modify_keywords:
            if keyword in text:
                return '修改'

        return '未知'

    def extract_changes_from_table(self, table, soup):
        """
        从表格中提取变更信息

        Args:
            table: BeautifulSoup表格元素
            soup: 完整的BeautifulSoup对象
        """
        rows = table.find_all('tr')
        if not rows or len(rows) < 2:
            return

        # 解析表头，确定列索引
        header_row = rows[0]
        headers = [th.get_text(strip=True) for th in header_row.find_all(['th', 'td'])]

        # 查找关键列的索引
        col_indices = {
            '编号': None,
            '接口名称': None,
            '名称': None,
            '变更类型': None,
            '变更描述': None,
            '影响说明': None,
            '描述': None,
            '说明': None,
            '路径': None
        }

        for i, header in enumerate(headers):
            for key in list(col_indices.keys()):
                if key in header or header in key:
                    col_indices[key] = i
                    # 找到后移除，避免重复匹配
                    if key in col_indices:
                        continue

        # 如果没有找到变更类型列，打印警告
        if col_indices['变更类型'] is None:
            print(f"  警告: 表格中未找到'变更类型'列，表头: {headers}")
            return

        print(f"  解析表格，表头: {headers}")

        # 获取表格所在章节
        table_chapter = self.find_chapter_for_element(table, soup)

        # 检查章节是否应该被包含
        if not self._should_include_chapter(table_chapter):
            print(f"  跳过章节: {table_chapter}")
            return

        print(f"  处理章节: {table_chapter}")

        # 解析数据行
        for row in rows[1:]:
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue

            # 提取各列数据
            values = [cell.get_text(strip=True) for cell in cells]

            # 获取变更类型
            change_type_col = col_indices['变更类型']
            if change_type_col is not None and change_type_col < len(values):
                change_type_raw = values[change_type_col]
                change_type = self.normalize_change_type(change_type_raw)
            else:
                continue

            # 如果变更类型未知，跳过
            if change_type == '未知':
                continue

            # 提取变更项名称（接口名称/名称）
            item_name = ''
            for name_key in ['接口名称', '名称']:
                if col_indices.get(name_key) is not None and col_indices[name_key] < len(values):
                    name = values[col_indices[name_key]]
                    if name:
                        item_name = name
                        break

            # 如果没有找到名称，尝试其他列
            if not item_name:
                # 尝试获取变更类型列之外的第一列非空数据
                for i, val in enumerate(values):
                    if i != change_type_col and val and col_indices['编号'] != i:
                        item_name = val
                        break

            # 提取描述信息（优先使用变更描述，其次使用描述）
            description = ''
            for desc_key in ['变更描述', '描述']:
                if col_indices.get(desc_key) is not None and col_indices[desc_key] < len(values):
                    desc = values[col_indices[desc_key]]
                    if desc:
                        description = desc
                        break

            # 如果没有找到描述，尝试使用路径列
            if not description and col_indices.get('路径') is not None and col_indices['路径'] < len(values):
                path = values[col_indices['路径']]
                if path:
                    description = path

            # 提取影响说明（单独列）
            impact = ''
            if col_indices.get('影响说明') is not None and col_indices['影响说明'] < len(values):
                impact = values[col_indices['影响说明']]

            self.changes.append({
                'chapter': table_chapter,
                'change_type': change_type,
                'item_name': item_name,
                'description': description,
                'impact': impact,
                'change_type_raw': change_type_raw,
                'verified': '待验证',
                'remark': ''
            })

    def extract_changes(self, soup):
        """
        提取所有变更信息

        Args:
            soup: BeautifulSoup对象
        """
        print("正在查找变更信息...")

        # 首先查找所有标题，记录章节信息
        all_headings = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        print(f"  找到 {len(all_headings)} 个章节标题")

        # 方法1: 查找所有带data-change属性的元素
        change_elements = soup.find_all(attrs={'data-change': True})
        if change_elements:
            print(f"  找到 {len(change_elements)} 个带data-change属性的元素")

        for elem in change_elements:
            change_type = elem.get('data-change')
            text = elem.get_text(strip=True)

            if not text:
                continue

            chapter_path = self.find_chapter_for_element(elem, soup)

            change_type_map = {
                'deleted': '删除',
                'added': '新增',
                'modified': '修改'
            }
            change_type_cn = change_type_map.get(change_type, change_type)

            self.changes.append({
                'chapter': chapter_path,
                'change_type': change_type_cn,
                'item_name': '',
                'description': text,
                'change_type_raw': change_type,
                'verified': '待验证',
                'remark': ''
            })

        # 方法2: 查找表格中的变更标记（使用CSS类）
        for tag_class, change_type in [
            ('tag-deleted', '删除'),
            ('tag-added', '新增'),
            ('tag-modified', '修改')
        ]:
            tags = soup.find_all(class_=tag_class)
            if tags:
                print(f"  找到 {len(tags)} 个带{tag_class}类的元素")

            for tag in tags:
                row = tag.find_parent('tr')
                if row:
                    cells = row.find_all('td')
                    if len(cells) >= 2:
                        text = cells[0].get_text(strip=True)
                        description = cells[3].get_text(strip=True) if len(cells) > 3 else ''
                        chapter_path = self.find_chapter_for_element(row, soup)

                        self.changes.append({
                            'chapter': chapter_path,
                            'change_type': change_type,
                            'item_name': text,
                            'description': description,
                            'change_type_raw': change_type.lower().replace('删除', 'deleted').replace('新增', 'added').replace('修改', 'modified'),
                            'verified': '待验证',
                            'remark': ''
                        })

        # 方法3: 查找包含变更类型关键词的表格
        tables = soup.find_all('table')
        print(f"  共找到 {len(tables)} 个表格")

        for table in tables:
            # 检查表格是否包含变更相关的表头
            table_text = table.get_text()
            change_keywords = ['变更类型', '变更', '删除', '新增', '修改', '接口', '目录']

            # 如果表格包含变更相关关键词，尝试解析
            if any(keyword in table_text for keyword in change_keywords):
                self.extract_changes_from_table(table, soup)

    def create_sheet(self, wb, sheet_name, changes, color):
        """
        创建一个sheet页

        Args:
            wb: 工作簿对象
            sheet_name: sheet名称
            changes: 变更数据列表
            color: 表头颜色
        """
        # 删除默认sheet（如果存在且不是我们要创建的）
        if 'Sheet' in wb.sheetnames and sheet_name not in wb.sheetnames:
            del wb['Sheet']

        # 创建sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 清空现有数据
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 定义样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 定义列头
        headers = ['章节', '变更项', '影响说明', '描述', '验证状态', '备注']

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = alignment

        # 写入数据
        for row_num, change in enumerate(changes, 2):
            ws.cell(row=row_num, column=1, value=change['chapter'])
            ws.cell(row=row_num, column=2, value=change['item_name'])
            ws.cell(row=row_num, column=3, value=change.get('impact', ''))
            ws.cell(row=row_num, column=4, value=change['description'])
            ws.cell(row=row_num, column=5, value=change['verified'])
            ws.cell(row=row_num, column=6, value=change['remark'])

            # 应用边框和对齐
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = alignment

        # 调整列宽
        ws.column_dimensions['A'].width = 35  # 章节
        ws.column_dimensions['B'].width = 30  # 变更项
        ws.column_dimensions['C'].width = 30  # 影响说明
        ws.column_dimensions['D'].width = 40  # 描述
        ws.column_dimensions['E'].width = 12  # 验证状态
        ws.column_dimensions['F'].width = 30  # 备注

        # 冻结首行
        ws.freeze_panes = 'A2'

    def export_to_excel(self):
        """将变更信息导出到Excel，按变更类型分sheet"""
        # 按变更类型分组
        changes_by_type = {
            '删除': [],
            '新增': [],
            '修改': []
        }

        for change in self.changes:
            change_type = change['change_type']
            if change_type in changes_by_type:
                changes_by_type[change_type].append(change)

        # 创建工作簿
        wb = Workbook()
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 定义各类型的颜色
        type_colors = {
            '删除': 'D32F2F',    # 红色
            '新增': '388E3C',    # 绿色
            '修改': 'F57C00'     # 橙色
        }

        # 创建各类型的sheet
        for change_type, color in type_colors.items():
            changes = changes_by_type[change_type]
            if changes:  # 只创建有数据的sheet
                self.create_sheet(wb, change_type, changes, color)

        # 确保输出目录存在
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)

        # 保存文件
        wb.save(self.output_file)

        # 打印各类型的统计
        print(f"\nExcel报告已生成: {self.output_file}")
        print("各Sheet页数据统计:")
        for change_type, changes in changes_by_type.items():
            print(f"  {change_type}: {len(changes)} 条")

    def generate_summary(self):
        """生成变更摘要"""
        summary = {
            '删除': 0,
            '新增': 0,
            '修改': 0,
            '总计': len(self.changes)
        }

        for change in self.changes:
            change_type = change['change_type']
            if change_type in summary:
                summary[change_type] += 1

        return summary

    def parse(self):
        """执行解析"""
        print(f"正在解析: {self.html_file}")
        soup = self.load_html()
        self.extract_changes(soup)

        if not self.changes:
            print("警告: 未找到任何变更标记")
            return

        # 打印摘要
        summary = self.generate_summary()
        print("\n变更摘要:")
        for change_type, count in summary.items():
            print(f"  {change_type}: {count}")

        # 导出到Excel
        self.export_to_excel()


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='HTML差异文档解析器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 解析所有章节
  python diff_parser.py file.html

  # 只解析第5章
  python diff_parser.py file.html --chapters 5

  # 只解析第5.1和5.2节
  python diff_parser.py file.html --chapters 5.1,5.2

  # 跳过第1和第2章
  python diff_parser.py file.html --skip 1,2

  # 只解析第5章，跳过5.1节
  python diff_parser.py file.html --chapters 5 --skip 5.1
        """
    )
    parser.add_argument('html_file', help='HTML差异文档路径')
    parser.add_argument('-o', '--output', help='输出Excel文件路径')
    parser.add_argument('-c', '--chapters', help='只解析指定章节，如: 5 或 5.1,5.2')
    parser.add_argument('-s', '--skip', help='跳过指定章节，如: 1,2 或 5.1')

    args = parser.parse_args()

    # 执行解析
    diff_parser = DiffParser(
        args.html_file,
        args.output,
        include_chapters=args.chapters,
        skip_chapters=args.skip
    )
    diff_parser.parse()


if __name__ == '__main__':
    main()
